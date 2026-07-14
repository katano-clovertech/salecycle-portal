"""
SaleCycle Daily Data Collector
Extracts Sends/Opens/Clicks/Conversions for all clients from Looker dashboards
and appends results to an Excel file.
"""
import os
import sys
import time
import json
import datetime
import pandas as pd
import requests as req_lib
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Configuration ---
EMAIL = os.environ.get("SALECYCLE_USER", "s.katano@clovertech.jp")
PASSWORD = os.environ.get("SALECYCLE_PASS", "")
SLACK_WEBHOOK_URL = os.environ.get("SLACK_WEBHOOK_URL", "")
GOOGLE_SHEETS_ID = "16i0Mwrsx0o8DwLfSRIRjDcEy_rbhdNbO9mTcnyZkpBw"
EXCEL_INPUT  = os.path.join(os.path.dirname(__file__), "salecycle動作確認.xlsx")
CLIENTS_CSV  = os.path.join(os.path.dirname(__file__), "clients.csv")
EXCEL_OUTPUT = os.path.join(os.path.dirname(__file__), "salecycle_daily_report.xlsx")
LOOKER_API_BASE = "https://looker-api.salecycle.com/api/internal"
MY_SALECYCLE_BASE = "https://my.salecycle.com"

# Date range: set dynamically in main() as absolute date (e.g. "2026-03-15 to 2026-03-15")
DATE_RANGE = ""
DATE_GRANULARITY = "Day"
CURRENCY = "JPY"

# Dashboard URLs
DASHBOARD_URLS = {
    "basket": f"{MY_SALECYCLE_BASE}/dashboard/new_business_aggregates::basket_abandonment__campaign_aggregates",
    "browse": f"{MY_SALECYCLE_BASE}/dashboard/new_business_aggregates::browse_abandonment__campaign_aggregates",
    "display": f"{MY_SALECYCLE_BASE}/dashboard/new_business_aggregates::display_only__campaign_aggregates",
}

# Metric fields in query results
METRIC_FIELDS = {
    "basket": {
        "sends": "campaign_aggregates.m_sends",
        "opens": "campaign_aggregates.m_opens",
        "clicks": "campaign_aggregates.m_clicks",
        "conversions": "campaign_aggregates.m_dispatch_conversions",
        "revenue": "campaign_aggregates.m_dispatch_conversion_value",
    },
    "browse": {
        "sends": "campaign_aggregates.m_sends",
        "opens": "campaign_aggregates.m_opens",
        "clicks": "campaign_aggregates.m_clicks",
        "conversions": "campaign_aggregates.m_dispatch_conversions",
        "revenue": "campaign_aggregates.m_dispatch_conversion_value",
    },
    "display": {
        "sends": "campaign_aggregates.m_displays",
        "opens": None,
        "clicks": "campaign_aggregates.m_display_clicks",
        "conversions": "campaign_aggregates.m_display_conversions",
        "revenue": "campaign_aggregates.m_display_conversion_value",
    },
}


def login_and_get_session(page):
    """Log in to my.salecycle.com"""
    print("Logging in to my.salecycle.com...")
    page.goto(MY_SALECYCLE_BASE)
    page.wait_for_load_state("networkidle", timeout=30000)
    time.sleep(2)
    page.fill('input[type="email"]', EMAIL)
    page.fill('input[type="password"]', PASSWORD)
    page.click('button:has-text("Sign in")')
    try:
        page.wait_for_url(lambda url: url != f"{MY_SALECYCLE_BASE}/", timeout=20000)
    except:
        pass
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except:
        pass
    time.sleep(3)
    print(f"Logged in: {page.url}")


def capture_all_templates(context, page, needed_dashboards):
    """Navigate to each dashboard and capture querymanager request bodies.
    Uses a single global listener. Requests fire ~90-120s after navigation,
    so we navigate all dashboards first, then wait for all captures."""
    templates = {}

    def on_request(req):
        if "querymanager/queries" not in req.url or req.method != "POST":
            return
        try:
            body_text = req.post_data
        except Exception:
            return
        if not body_text:
            print(f"  [qm] querymanager POST with no body")
            return
        try:
            body = json.loads(body_text)
            ctx = body.get("context", {}).get("id", "")
            print(f"  [qm] ctx={ctx}")
            for dtype in ["basket", "browse", "display"]:
                if dtype in ctx and dtype not in templates:
                    templates[dtype] = body
                    print(f"  Captured {dtype} template (ctx={ctx})")
        except Exception as e:
            print(f"  [qm] error: {e}")


    context.on("request", on_request)

    # Navigate to each dashboard to trigger Looker embed loading
    for dash_type in sorted(needed_dashboards):
        print(f"Loading {dash_type} dashboard...")
        page.goto(DASHBOARD_URLS[dash_type])
        try:
            page.wait_for_selector("iframe:not([aria-hidden])", timeout=60000)
        except Exception:
            pass
        time.sleep(10)  # Brief pause before next navigation

    # Wait up to 5 minutes total for all templates to be captured
    print("Waiting for all dashboard templates (up to 5 min)...")
    deadline = time.time() + 300
    last_count = 0
    while time.time() < deadline and len(templates) < len(needed_dashboards):
        time.sleep(5)
        if len(templates) > last_count:
            print(f"  Templates so far: {sorted(templates.keys())}")
            last_count = len(templates)

    context.remove_listener("request", on_request)

    for dtype in sorted(needed_dashboards):
        if dtype in templates:
            print(f"  {dtype}: captured ({len(templates[dtype].get('saved_queries', []))} queries)")
        else:
            print(f"  WARNING: No template for {dtype}")
    return templates


def extract_field_value(row, field_name):
    """Extract numeric value from a row field, handling pivoted data structure."""
    val = row.get(field_name)
    if val is None:
        return 0
    if isinstance(val, dict):
        # Direct value (non-pivoted)
        if "value" in val:
            return val.get("value") or 0
        # Pivoted: {pivot_key: {"value": X}, ...} — sum across all pivot buckets
        total = 0
        for pivot_val in val.values():
            if isinstance(pivot_val, dict):
                total += pivot_val.get("value") or 0
        return total
    return val or 0


def extract_metrics_from_result(result_data, dashboard_type):
    """Extract and sum metrics from a complete query result dict."""
    fields = METRIC_FIELDS[dashboard_type]
    totals = {k: 0 for k in fields}
    rows = (result_data.get("data") or {}).get("data") or []
    for field_key, field_name in fields.items():
        if field_name:
            total = sum(extract_field_value(row, field_name) for row in rows)
            if total > 0:
                totals[field_key] = total
    return totals


def parse_ndjson_response(text, dashboard_type):
    """Parse streaming NDJSON response and sum metrics across all rows"""
    fields = METRIC_FIELDS[dashboard_type]
    totals = {k: 0 for k in fields}

    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        try:
            data = json.loads(line)
        except:
            continue

        if data.get("status") != "complete":
            continue

        partial = extract_metrics_from_result(data, dashboard_type)
        for k, v in partial.items():
            if v > 0:
                totals[k] = v

    return totals


def fetch_metrics_for_client(session, headers, base_body, client_name, dashboard_type):
    """Make API call for a specific client and extract metrics"""
    if not base_body:
        return None

    import copy
    modified_body = copy.deepcopy(base_body)
    if "options" in modified_body:
        modified_body["options"]["force_run"] = True

    # Filter to only the element with date-based sorts (main metrics chart)
    all_sqs = modified_body.get("saved_queries", [])
    date_sqs = [sq for sq in all_sqs
                if any("date" in s for s in sq.get("sorts", []))]
    if date_sqs:
        modified_body["saved_queries"] = [date_sqs[0]]
    elif all_sqs:
        modified_body["saved_queries"] = [all_sqs[0]]

    for sq in modified_body.get("saved_queries", []):
        for f in sq.get("filters", []):
            if "new_clients.client_name" in f:
                f["new_clients.client_name"] = client_name
            if "new_currency_exchange_rates.currency_exchange" in f:
                f["new_currency_exchange_rates.currency_exchange"] = CURRENCY
            if "campaign_aggregates.time_slice" in f:
                f["campaign_aggregates.time_slice"] = DATE_RANGE
            if "campaign_aggregates.date_granularity" in f:
                f["campaign_aggregates.date_granularity"] = DATE_GRANULARITY
            if "new_business_aggregates.time_slice" in f:
                f["new_business_aggregates.time_slice"] = DATE_RANGE
            if "new_business_aggregates.date_granularity" in f:
                f["new_business_aggregates.date_granularity"] = DATE_GRANULARITY

    try:
        resp = session.post(
            f"{LOOKER_API_BASE}/querymanager/queries",
            headers=headers,
            json=modified_body,
            timeout=30
        )
        # Update CSRF token from response cookies
        for c in resp.cookies:
            if c.name == "CSRF-TOKEN":
                headers["X-CSRF-Token"] = c.value
        if resp.status_code != 200:
            print(f"    API error {resp.status_code}: {resp.text[:100]}")
            return None
    except Exception as e:
        print(f"    Request error: {e}")
        return None

    # Parse initial response lines; collect pending query IDs
    get_hdrs = {k: v for k, v in headers.items() if k != "Content-Type"}
    totals = {k: 0 for k in METRIC_FIELDS[dashboard_type]}

    for line in resp.text.split('\n'):
        line = line.strip()
        if not line:
            continue
        try:
            item = json.loads(line)
        except:
            continue

        if item.get("status") == "complete":
            partial = extract_metrics_from_result(item, dashboard_type)
            for k, v in partial.items():
                if v > 0:
                    totals[k] = v
            continue

        qid = item.get("id")
        if not qid:
            continue

        # Poll until complete
        deadline = time.time() + 90
        while time.time() < deadline:
            time.sleep(2)
            try:
                pr = session.get(
                    f"{LOOKER_API_BASE}/querymanager/queries/{qid}",
                    headers=get_hdrs, timeout=30
                )
                if pr.status_code == 200:
                    pd_data = pr.json()
                    if isinstance(pd_data, dict) and pd_data.get("status") == "complete":
                        partial = extract_metrics_from_result(pd_data, dashboard_type)
                        for k, v in partial.items():
                            if v > 0:
                                totals[k] = v
                        break
            except Exception as e:
                print(f"    Poll error: {e}")

    return totals


def read_clients_from_excel():
    """Read client list from clients.csv (preferred) or Excel file"""
    # クラウド実行 or ローカルでCSVがあればCSVを使用
    if os.path.exists(CLIENTS_CSV):
        df = pd.read_csv(CLIENTS_CSV, encoding="utf-8-sig")
        clients = []
        for _, row in df.iterrows():
            name = str(row["client_name"]).strip()
            if not name or name == "nan":
                continue
            dashboards = []
            def _has(col):
                v = row.get(col)
                try:
                    return pd.notna(v) and float(v) == 1.0
                except (ValueError, TypeError):
                    return False
            if _has("basket"):
                dashboards.append("basket")
            if _has("browse"):
                dashboards.append("browse")
            if _has("display"):
                dashboards.append("display")
            if dashboards:
                clients.append({"name": name, "dashboards": dashboards})
        print(f"Loaded {len(clients)} clients from clients.csv")
        return clients

    # フォールバック: Excelから読み込み
    df = pd.read_excel(EXCEL_INPUT, header=1)
    clients = []
    for _, row in df.iterrows():
        name = str(row.iloc[0]).strip()
        if not name or name == "nan" or name == "クライアント":
            continue
        basket_url  = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        browse_url  = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
        display_url = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""

        dashboards = []
        if basket_url and basket_url != "nan":
            dashboards.append("basket")
        if browse_url and browse_url != "nan":
            dashboards.append("browse")
        if display_url and display_url != "nan":
            dashboards.append("display")

        if dashboards:
            clients.append({"name": name, "dashboards": dashboards})

    return clients


def save_to_excel(results, report_date):
    """Save or append results to Excel report (deduplicates by date+client+dashboard)"""
    dashboard_labels = {"basket": "Basket", "browse": "Browse", "display": "Display"}

    try:
        wb = load_workbook(EXCEL_OUTPUT)
        ws = wb.active
        # Build set of already-written (date, client, dashboard) keys to avoid duplicates
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                existing.add((str(row[0]), str(row[1]), str(row[2])))
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        headers = ["日付", "クライアント", "ダッシュボード種別", "送付件数", "開封数", "クリック数", "コンバージョン数", "コンバージョン金額"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = PatternFill("solid", start_color="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.alignment = Alignment(horizontal="center")
        existing = set()

    next_row = ws.max_row + 1
    added = 0

    for item in results:
        label = dashboard_labels.get(item["dashboard"], item["dashboard"])
        key = (str(report_date), item["client"], label)
        if key in existing:
            continue  # Skip duplicate
        existing.add(key)
        row_data = [
            report_date,
            item["client"],
            label,
            item.get("sends", 0),
            item.get("opens", "") if item.get("opens") is not None else "",
            item.get("clicks", 0),
            item.get("conversions", 0),
            item.get("revenue", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center")
        next_row += 1
        added += 1

    print(f"Saving {added} new rows to Excel...")

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(EXCEL_OUTPUT)
    print(f"Saved to {EXCEL_OUTPUT}")

    # CSVに保存
    try:
        import pandas as _pd, subprocess as _sp
        _csv_path = os.path.join(os.path.dirname(__file__), "salecycle_daily_report.csv")

        if os.environ.get("HEADLESS", "0") == "1":
            # クラウド実行: 既存CSVに追記（Excelなしでも過去データを保持）
            new_rows = []
            for item in results:
                _label = {"basket": "Basket", "browse": "Browse", "display": "Display"}.get(item["dashboard"], item["dashboard"])
                new_rows.append({
                    "日付": report_date,
                    "クライアント": item["client"],
                    "ダッシュボード種別": _label,
                    "送付件数": item.get("sends", 0),
                    "開封数": item.get("opens", ""),
                    "クリック数": item.get("clicks", 0),
                    "コンバージョン数": item.get("conversions", 0),
                    "コンバージョン金額": item.get("revenue", ""),
                })
            if new_rows:
                _new_df = _pd.DataFrame(new_rows)
                if os.path.exists(_csv_path):
                    _existing = _pd.read_csv(_csv_path, encoding="utf-8-sig")
                    # 重複排除してマージ
                    _combined = _pd.concat([_existing, _new_df], ignore_index=True)
                    _combined = _combined.drop_duplicates(subset=["日付", "クライアント", "ダッシュボード種別"], keep="last")
                else:
                    _combined = _new_df
                _combined.to_csv(_csv_path, index=False, encoding="utf-8-sig")
                print(f"CSV updated (cloud): {len(_combined)} total rows")
        else:
            # ローカル実行: Excelから全データでCSV生成してgit push
            _df = _pd.read_excel(EXCEL_OUTPUT, engine="openpyxl")
            _df.to_csv(_csv_path, index=False, encoding="utf-8-sig")
            print(f"CSV saved: {_csv_path}")
            _repo = os.path.dirname(__file__)
            _sp.run(["git", "-C", _repo, "add", "salecycle_daily_report.csv"], check=True)
            _sp.run(["git", "-C", _repo, "commit", "-m", f"data: {report_date}"], check=True)
            _sp.run(["git", "-C", _repo, "push"], check=True)
            print(f"CSV pushed to GitHub ({report_date})")
    except Exception as _e:
        print(f"CSV save error: {_e}")


def get_previous_sends(report_date):
    """前日の送付件数を {(client, dashboard_label): sends} で返す。
    クラウド実行はCSV、ローカル実行はExcelから読み込む。"""
    prev_date = (datetime.datetime.strptime(report_date, "%Y-%m-%d") - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    prev_sends = {}
    _csv_path = os.path.join(os.path.dirname(__file__), "salecycle_daily_report.csv")

    # クラウド実行 or Excelがない場合はCSVから読む
    if os.environ.get("HEADLESS", "0") == "1" or not os.path.exists(EXCEL_OUTPUT):
        if os.path.exists(_csv_path):
            try:
                df = pd.read_csv(_csv_path, encoding="utf-8-sig")
                prev_rows = df[df["日付"].astype(str) == prev_date]
                for _, row in prev_rows.iterrows():
                    key = (row["クライアント"], row["ダッシュボード種別"])
                    sends = row["送付件数"]
                    prev_sends[key] = sends if isinstance(sends, (int, float)) else 0
                print(f"  [Slack] 前日データ({prev_date}): {len(prev_sends)}件 (CSV)")
            except Exception as e:
                print(f"  [Slack] 前日データ読み込みエラー(CSV): {e}")
        else:
            print(f"  [Slack] CSVが見つからないため前日比チェックをスキップ")
        return prev_sends

    # ローカル実行: Excelから読む
    try:
        wb = load_workbook(EXCEL_OUTPUT, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val, client, dashboard, sends = row[0], row[1], row[2], row[3]
            if str(date_val) == prev_date and client and dashboard:
                prev_sends[(client, dashboard)] = sends if isinstance(sends, (int, float)) else 0
        wb.close()
        print(f"  [Slack] 前日データ({prev_date}): {len(prev_sends)}件 (Excel)")
    except Exception as e:
        print(f"  [Slack] 前日データ読み込みエラー: {e}")
    return prev_sends


def send_slack_report(alerts, results, report_date):
    """日次レポートをSlackに送信する（アラートの有無に関わらず常に送信）"""
    if not SLACK_WEBHOOK_URL:
        print("  [Slack] SLACK_WEBHOOK_URL \u304c\u672a\u8a2d\u5b9a\u306e\u305f\u3081\u901a\u77e5\u3092\u30b9\u30ad\u30c3\u30d7")
        return

    total_sends = sum(
        int(item["sends"]) for item in results
        if isinstance(item.get("sends"), (int, float))
    )
    client_count = len(set(item["client"] for item in results))

    excel_path = EXCEL_OUTPUT.replace("\\", "/")
    excel_link = f"file:///{excel_path}"

    lines = [
        f":bar_chart: *SaleCycle \u65e5\u6b21\u30ec\u30dd\u30fc\u30c8 ({report_date})*",
        f"\u51e6\u7406\u30af\u30e9\u30a4\u30a2\u30f3\u30c8\u6570: {client_count}\u4ef6 | \u5408\u8a08\u9001\u4ed8\u4ef6\u6570: {total_sends:,}\u4ef6",
        "",
    ]

    # アラート種別カウント
    total_rows = len(results)
    n_red    = sum(1 for a in alerts if a["reason"] in ("fetch_failed", "zero"))
    n_yellow = sum(1 for a in alerts if a["reason"] == "drop" and a["change_pct"] <= -50)
    n_green  = sum(1 for a in alerts if a["reason"] == "drop" and a["change_pct"] > -50)
    n_white  = total_rows - len(alerts)

    summary = (
        f":red_circle: {n_red}件  "
        f":large_yellow_circle: {n_yellow}件  "
        f":large_green_circle: {n_green}件  "
        f":white_circle: {n_white}件"
    )
    lines.append(summary)
    lines.append("")


    if alerts:
        # 赤→黄→緑の順にソート
        def _alert_order(a):
            if a["reason"] in ("fetch_failed", "zero"): return 0
            return 1 if a["change_pct"] <= -50 else 2
        sorted_alerts = sorted(alerts, key=_alert_order)

        lines.append(f":warning: *アラート {len(alerts)}件:*")
        lines.append(":red_circle: 0件・取得失敗　:large_yellow_circle: 50%以上減　:large_green_circle: 20%以上減")
        for a in sorted_alerts:
            if a["reason"] == "fetch_failed":
                lines.append(f"- :red_circle: {a['client']} [{a['dashboard']}]: *データ取得失敗*（前日: {a['prev']:,}件）")
            elif a["reason"] == "zero":
                lines.append(f"- :red_circle: {a['client']} [{a['dashboard']}]: *0件*（前日: {a['prev']:,}件）")
            else:
                pct = a["change_pct"]
                icon = ":large_yellow_circle:" if pct <= -50 else ":large_green_circle:"
                lines.append(
                    f"- {icon} {a['client']} [{a['dashboard']}]: {a['today']:,}件"
                    f"（前日: {a['prev']:,}件 / {pct:+.1f}%）"
                )
    # ⚪ 問題なし一覧
    _dash_labels = {"basket": "Basket", "browse": "Browse", "display": "Display"}
    alert_keys = {(a["client"], a["dashboard"]) for a in alerts}
    white_items = [
        item for item in results
        if (item["client"], _dash_labels.get(item["dashboard"], item["dashboard"])) not in alert_keys
        and isinstance(item.get("sends"), (int, float))
    ]
    if white_items:
        lines.append(f":white_circle: *問題なし {len(white_items)}件:*")
        for item in white_items:
            label = _dash_labels.get(item["dashboard"], item["dashboard"])
            sends = int(item["sends"])
            lines.append(f"- :white_circle: {item['client']} [{label}]: {sends:,}件")
    elif not alerts:
        lines.append(":white_check_mark: 異常なし")

    lines.append("")
    lines.append(f":open_file_folder: <{excel_link}|Excel\u30ec\u30dd\u30fc\u30c8\u3092\u958b\u304f>")
    lines.append(f":bar_chart: <https://salecycle-portal.streamlit.app/|\u30dd\u30fc\u30bf\u30eb\u30b5\u30a4\u30c8\u3092\u958b\u304f>")

    payload = {"text": "\n".join(lines)}
    try:
        resp = req_lib.post(SLACK_WEBHOOK_URL, json=payload, timeout=10)
        if resp.status_code == 200:
            alert_msg = f"{len(alerts)}\u4ef6\u306e\u30a2\u30e9\u30fc\u30c8\u3042\u308a" if alerts else "\u7570\u5e38\u306a\u3057"
            print(f"  [Slack] \u30ec\u30dd\u30fc\u30c8\u3092\u9001\u4fe1\u3057\u307e\u3057\u305f ({alert_msg})")
        else:
            print(f"  [Slack] \u9001\u4fe1\u5931\u6557: {resp.status_code} {resp.text[:80]}")
    except Exception as e:
        print(f"  [Slack] \u9001\u4fe1\u30a8\u30e9\u30fc: {e}")


def check_sends_alerts(results, report_date):
    update_google_sheets(results, report_date)
    """\u9001\u4ed8\u4ef6\u6570\u30c1\u30a7\u30c3\u30af\u30fb\u65e5\u6b21\u30ec\u30dd\u30fc\u30c8\u3092Slack\u306b\u9001\u4fe1\u3059\u308b"""
    dashboard_labels = {"basket": "Basket", "browse": "Browse", "display": "Display"}
    prev_sends = get_previous_sends(report_date)
    alerts = []

    for item in results:
        today_sends = item.get("sends")
        label = dashboard_labels.get(item["dashboard"], item["dashboard"])
        key = (item["client"], label)
        prev = prev_sends.get(key)

        if today_sends == "" or today_sends is None:
            # 取得失敗
            alerts.append({
                "client": item["client"], "dashboard": label,
                "today": None, "prev": prev if prev is not None else 0,
                "reason": "fetch_failed", "change_pct": None,
            })
            continue

        if not isinstance(today_sends, (int, float)):
            continue
        today_sends = int(today_sends)

        if today_sends == 0:
            alerts.append({
                "client": item["client"], "dashboard": label,
                "today": 0, "prev": prev if prev is not None else 0,
                "reason": "zero", "change_pct": -100.0,
            })
        elif prev is not None and prev > 0:
            change_pct = (today_sends - prev) / prev * 100
            if change_pct <= -20:
                alerts.append({
                    "client": item["client"], "dashboard": label,
                    "today": today_sends, "prev": int(prev),
                    "reason": "drop", "change_pct": change_pct,
                })

    if alerts:
        print(f"\nSlack\u30a2\u30e9\u30fc\u30c8\u5bfe\u8c61: {len(alerts)} \u4ef6")
        for a in alerts:
            if a["reason"] == "fetch_failed":
                reason = "取得失敗"
            elif a["reason"] == "zero":
                reason = "0件"
            else:
                reason = f"{a['change_pct']:+.1f}%"
            print(f"  {a['client']} [{a['dashboard']}]: {reason}")
    else:
        print("\nSlack\u30a2\u30e9\u30fc\u30c8: \u7570\u5e38\u306a\u3057")

    send_slack_report(alerts, results, report_date)



def update_google_sheets(results, report_date):
    """グーグルスプレッドシートに当日データを書き込む（再実行時は上書き）"""
    import json as _json
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not creds_json:
        print("  [Sheets] GOOGLE_SERVICE_ACCOUNT_JSON が未設定のためスキップ")
        return
    try:
        import gspread
        from google.oauth2.service_account import Credentials as _Creds
        creds = _Creds.from_service_account_info(
            _json.loads(creds_json),
            scopes=["https://www.googleapis.com/auth/spreadsheets"],
        )
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(GOOGLE_SHEETS_ID).sheet1

        # ヘッダー確認
        all_vals = ws.get_all_values()
        header = ["日付", "クライアント", "ダッシュボード種別", "送付件数", "開封数", "クリック数", "コンバージョン数", "コンバージョン金額"]
        if not all_vals:
            ws.append_row(header)
            all_vals = [header]

        # 今日の既存行を削除（再実行時の重複防止）
        today_rows = [i + 1 for i, row in enumerate(all_vals) if i > 0 and row and row[0] == report_date]
        for idx in sorted(today_rows, reverse=True):
            ws.delete_rows(idx)

        # 新データ追記
        dash_labels = {"basket": "Basket", "browse": "Browse", "display": "Display"}
        new_rows = [
            [report_date, item["client"], dash_labels.get(item["dashboard"], item["dashboard"]),
             item.get("sends", ""), item.get("opens", ""), item.get("clicks", ""),
             item.get("conversions", ""), item.get("revenue", "")]
            for item in results
        ]
        if new_rows:
            ws.append_rows(new_rows, value_input_option="USER_ENTERED")
        print(f"  [Sheets] {len(new_rows)}行を書き込みました ({report_date})")
    except Exception as e:
        print(f"  [Sheets] 更新エラー: {e}")

def find_missing_dates(days_back=7):
    """過去days_back日間でExcelにデータがない日付を (days_ago, date_str) のリストで返す。
    16時前は昨日（days_ago=1）をスキップ: PM3:00更新前のため不完全データになるのを防ぐ。
    16時以降は昨日も含める（16時タスクが失敗した場合のフォールバック）。
    """
    now = datetime.datetime.now()
    today = now.date()
    # 16:00より前は昨日をスキップ（16時タスクに任せる）
    min_days_ago = 2 if now.hour < 16 else 1

    existing_dates = set()

    if os.path.exists(EXCEL_OUTPUT):
        try:
            wb = load_workbook(EXCEL_OUTPUT, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    existing_dates.add(str(row[0])[:10])  # YYYY-MM-DD
            wb.close()
        except Exception as e:
            print(f"  [Backfill] Excel読み込みエラー: {e}")

    missing = []
    for days_ago in range(min_days_ago, days_back + 1):
        target = today - datetime.timedelta(days=days_ago)
        target_str = target.strftime("%Y-%m-%d")
        if target_str not in existing_dates:
            missing.append((days_ago, target_str))

    return missing


def load_templates_from_files():
    """Load dashboard request templates from pre-captured JSON files."""
    templates = {}
    script_dir = os.path.dirname(os.path.abspath(__file__))
    for dtype, fname in [("basket", "basket_req.json"), ("browse", "browse_req.json"),
                         ("display", "display_req.json")]:
        path = os.path.join(script_dir, fname)
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                templates[dtype] = json.load(f)
            sqs = len(templates[dtype].get("saved_queries", []))
            print(f"  Loaded {dtype} template ({sqs} queries)")
        else:
            print(f"  WARNING: {fname} not found")
    return templates


def get_looker_session():
    """ブラウザでログインしてLookerセッション（requests.Session, headers）を返す"""
    looker_cookies = {}
    with sync_playwright() as p:
        _headless = os.environ.get("HEADLESS", "0") == "1"
        _channel  = None if _headless else "chrome"
        browser = p.chromium.launch(headless=_headless, channel=_channel)
        context = browser.new_context()
        page = context.new_page()

        login_and_get_session(page)

        print("Establishing Looker session...")
        page.goto(DASHBOARD_URLS["basket"])
        try:
            page.wait_for_selector("iframe:not([aria-hidden])", timeout=60000)
        except Exception:
            pass
        time.sleep(15)

        cookies = context.cookies()
        looker_cookies = {c["name"]: c["value"] for c in cookies
                          if "looker-api.salecycle.com" in c.get("domain", "")}
        browser.close()

    if not looker_cookies:
        print("ERROR: No Looker session cookies obtained")
        sys.exit(1)

    print(f"Looker cookies: {list(looker_cookies.keys())}")
    session = req_lib.Session()
    session.cookies.update(looker_cookies)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "X-CSRF-Token": looker_cookies.get("CSRF-TOKEN", ""),
        "Origin": "https://looker-api.salecycle.com",
        "Referer": "https://looker-api.salecycle.com/",
    }
    return session, headers


def collect_for_date(session, headers, templates, clients, report_date, days_ago, skip_slack=False):
    """指定日のデータを収集してExcel保存・Slack通知する"""
    global DATE_RANGE
    if days_ago is None:
        today = datetime.datetime.now().date()
        target = datetime.datetime.strptime(report_date, "%Y-%m-%d").date()
        days_ago = (today - target).days
    DATE_RANGE = f"{days_ago} day{'s' if days_ago != 1 else ''} ago for 1 day"
    print(f"\n--- Collecting: {report_date} (filter: {DATE_RANGE}) ---")

    results = []
    for client in clients:
        client_name = client["name"]
        print(f"\n  {client_name}:")

        for dash_type in client["dashboards"]:
            template = templates.get(dash_type)
            if not template:
                print(f"    {dash_type}: no template available, skipping")
                continue

            metrics = fetch_metrics_for_client(session, headers, template, client_name, dash_type)
            if metrics:
                revenue_val = metrics.get("revenue")
                result = {
                    "client": client_name,
                    "dashboard": dash_type,
                    "sends": int(metrics.get("sends", 0)),
                    "opens": int(metrics.get("opens", 0)) if metrics.get("opens") is not None else "",
                    "clicks": int(metrics.get("clicks", 0)),
                    "conversions": int(metrics.get("conversions", 0)),
                    "revenue": int(revenue_val) if isinstance(revenue_val, (int, float)) and revenue_val else "",
                }
                results.append(result)
                print(f"    {dash_type}: Sends={result['sends']}, Opens={result['opens']}, Clicks={result['clicks']}, Conv={result['conversions']}, Revenue={result['revenue']}")
            else:
                print(f"    {dash_type}: failed to get data")
                results.append({
                    "client": client_name,
                    "dashboard": dash_type,
                    "sends": "",
                    "opens": "",
                    "clicks": "",
                    "conversions": "",
                    "revenue": "",
                })

    save_to_excel(results, report_date)
    if not skip_slack:
        check_sends_alerts(results, report_date)
    return results


def backfill_from_date(from_date_str):
    """カスタムバックフィル: 指定日から昨日までの欠損データを補完する"""
    print(f"=== Custom Backfill: {from_date_str} 〜 昨日 ===")

    today = datetime.datetime.now().date()
    start = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date()
    yesterday = today - datetime.timedelta(days=1)

    # 既存データを確認
    existing_dates = set()
    if os.path.exists(EXCEL_OUTPUT):
        try:
            wb = load_workbook(EXCEL_OUTPUT, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    existing_dates.add(str(row[0])[:10])
            wb.close()
        except Exception as e:
            print(f"  Excel読み込みエラー: {e}")

    # 欠損日付リストを作成
    missing = []
    d = start
    while d <= yesterday:
        d_str = d.strftime("%Y-%m-%d")
        if d_str not in existing_dates:
            days_ago = (today - d).days
            missing.append((days_ago, d_str))
        d += datetime.timedelta(days=1)

    if not missing:
        print("欠損データなし - バックフィル不要")
        return

    print(f"対象日付: {len(missing)}日分 ({missing[0][1]} 〜 {missing[-1][1]})")

    clients = read_clients_from_excel()
    print(f"Clients: {len(clients)}")

    needed_dashboards = set()
    for c in clients:
        needed_dashboards.update(c["dashboards"])

    print("\nLoading dashboard templates...")
    templates = load_templates_from_files()
    for dtype in sorted(needed_dashboards):
        if dtype not in templates:
            print(f"  ERROR: No template for '{dtype}'.")

    session, headers = get_looker_session()

    for days_ago, report_date in missing:
        collect_for_date(session, headers, templates, clients, report_date, days_ago, skip_slack=True)

    print("\n=== Custom Backfill Complete ===")


def startup_backfill():
    """起動時バックフィルモード: 過去7日間の欠損データを補完する"""
    print("=== Startup Backfill Mode ===")

    missing = find_missing_dates(days_back=14)
    if not missing:
        print("欠損データなし - バックフィル不要")
        return

    print(f"欠損日付: {[d for _, d in missing]}")

    clients = read_clients_from_excel()
    print(f"Clients to process: {len(clients)}")

    needed_dashboards = set()
    for c in clients:
        needed_dashboards.update(c["dashboards"])

    print("\nLoading dashboard templates...")
    templates = load_templates_from_files()
    for dtype in sorted(needed_dashboards):
        if dtype not in templates:
            print(f"  ERROR: No template for '{dtype}'.")

    session, headers = get_looker_session()

    for days_ago, report_date in missing:
        collect_for_date(session, headers, templates, clients, report_date, days_ago)

    print("\n=== Backfill Complete ===")


def main():
    """通常モード: 昨日のデータを収集する"""
    days_ago = 1
    report_date = (datetime.datetime.now() - datetime.timedelta(days=days_ago)).strftime("%Y-%m-%d")
    print(f"Collecting data for: {report_date}")

    clients = read_clients_from_excel()
    print(f"Clients to process: {len(clients)}")

    needed_dashboards = set()
    for c in clients:
        needed_dashboards.update(c["dashboards"])
    print(f"Dashboard types needed: {needed_dashboards}")

    print("\nLoading dashboard templates...")
    templates = load_templates_from_files()
    for dtype in sorted(needed_dashboards):
        if dtype not in templates:
            print(f"  ERROR: No template for '{dtype}'. Run capture_templates.py first.")

    session, headers = get_looker_session()
    collect_for_date(session, headers, templates, clients, report_date, days_ago)

    print("Done!")


def send_slack_error(error_msg, mode="main"):
    """スクリプトがエラーで落ちた時にSlackへ通知する"""
    if not SLACK_WEBHOOK_URL:
        return
    import traceback
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = (
        f":rotating_light: *SaleCycle \u30b9\u30af\u30ea\u30d7\u30c8\u30a8\u30e9\u30fc* ({now})\n"
        f"\u30e2\u30fc\u30c9: `{mode}`\n"
        f"```{error_msg[:800]}```"
    )
    try:
        req_lib.post(SLACK_WEBHOOK_URL, json={"text": text}, timeout=10)
    except Exception:
        pass


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="SaleCycle Daily Data Collector")
    parser.add_argument("--startup", action="store_true",
                        help="起動時バックフィルモード: 過去7日間の欠損データを補完")
    parser.add_argument("--from-date", metavar="YYYY-MM-DD",
                        help="指定日から昨日までの欠損データを一括収集")
    args = parser.parse_args()

    if not PASSWORD:
        print("ERROR: SALECYCLE_PASS environment variable not set")
        sys.exit(1)

    if args.from_date:
        mode = "backfill"
    elif args.startup:
        mode = "startup"
    else:
        mode = "main"

    try:
        if args.from_date:
            backfill_from_date(args.from_date)
        elif args.startup:
            startup_backfill()
        else:
            main()
    except Exception as e:
        import traceback
        err = traceback.format_exc()
        print(f"FATAL ERROR:\n{err}")
        send_slack_error(err, mode=mode)
        sys.exit(1)
